"""
Excel export.

Takes the same column/row data produced for a preview and writes a styled
.xlsx file (friendly headers, bold header row, frozen header, auto-sized
columns) to an in-memory buffer for download.
"""
import io
from typing import Any, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def build_workbook(
    display_names: List[str],
    rows: List[List[Any]],
    sheet_name: str = "Report",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")
    header_align = Alignment(vertical="center")

    # Header row with friendly names.
    ws.append(display_names)
    for col_idx, _ in enumerate(display_names, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows.
    for row in rows:
        ws.append(row)

    # Freeze the header and add an auto-filter.
    ws.freeze_panes = "A2"
    if display_names:
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(display_names))}{len(rows) + 1}"
        )

    # Auto-size columns based on the longest value (capped for readability).
    for col_idx, name in enumerate(display_names, start=1):
        max_len = len(str(name))
        for row in rows:
            if col_idx - 1 < len(row) and row[col_idx - 1] is not None:
                max_len = max(max_len, len(str(row[col_idx - 1])))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
